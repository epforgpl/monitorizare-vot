from openpyxl import Workbook, load_workbook
import sys

# Reads forms definitions from xlsx files and produce corresponding SQL to import into the database.

if len(sys.argv) < 2:
    print('You need to provide a xlsx file to open (and optionally a file to write sql to)')
    exit(1)

fin = sys.argv[1]
if len(sys.argv) > 2:
    fout = sys.argv[2]
else:
    parts = fin.split('.')
    fout = '.'.join(parts[:-1]) + '.sql'

out = open(fout, 'w', buffering=1)

wb = load_workbook(fin)  # type: Workbook # guess_types?

out.write("DELETE FROM FormVersions;\n")
out.write("DELETE FROM OptionsToQuestions;\n")
out.write("DELETE FROM Questions;\n")
out.write("DELETE FROM Options;\n")
out.write("DELETE FROM FormSections;\n")

out.flush()

options = {
}
option_id = -1
option_id += 1
out.write("INSERT INTO Options (Id, IsFreeText, Text, Hint) "
          + "VALUES ({}, {}, '{}', null);\n".format(option_id, 1, 'Kliknij i wpisz'))
# Text for number/time: Click and enter answer; until
FREE_OPTION_ID = option_id

question_id = -1
TYPES = {
    'Multiple': 0,
    'Single': 1,
    'Number': 2,
    'Time': 2,
    'Text': 2,
}

form_section_id = -1
o2o_id = -1
form_versions = {}

for sheet_name in wb.sheetnames:
    if sheet_name[0] == '~':
        if sheet_name[1:] == 'Versions':
            s = wb[sheet_name]
            for form in s.iter_rows(max_col=2):
                form_code, version = [v.value for v in form]
                if form_code is None:
                    continue
                form_versions[form_code.strip()] = version

        continue

    form_id, form_name = sheet_name.split('.', 2)
    form_name = form_name.strip()

    try:
        version = form_versions[form_id]
    except KeyError:
        raise KeyError("Please define version of the form {} in ~Versions tab.".format(form_id))
    # TODO Insert Descriptions when https://github.com/code4romania/monitorizare-vot/issues/67 gets implemented
    out.write("\n\nINSERT INTO FormVersions (Code, CurrentVersion) VALUES ('{}', {});\n".format(form_id, version))

    s = wb[sheet_name]

    # Sections are not used in clients, but group questions in the admin dashboard
    form_section_id += 1
    out.write("INSERT INTO FormSections (Id, Code, Description) VALUES ({}, '{}','not used');\n"
              .format(form_section_id, form_section_id))

    q_num = 0
    for q in s.iter_rows(min_row=2, max_col=6):
        question, field_type, values, alarm_values, more_option, more_hint = [v.value.strip() if v.value else v.value for v in q]
        hint = ''  # TODO what is it?

        if question is None:
            continue

        try:
            question_type = TYPES[field_type]
        except KeyError:
            raise KeyError("Missing question type in form {}, question: {}".format(form_id, question))

        question_id += 1
        q_num += 1
        out.write("\nINSERT INTO Questions (Id, FormCode, Code, IdSection, QuestionType, Text, Hint) "
                  + "VALUES ({}, '{}', '{}', {}, {}, '{}', '{}');\n"
                  .format(question_id, form_id, "{}.{:02d}".format(form_id, q_num), form_section_id, question_type, question.replace("'", "\\'"), hint))

        if values:
            values = [v.strip() for v in values.lower().split('/')]

            flagged = []
            if alarm_values:
                for v in alarm_values.lower().split('/'):
                    v = v.strip()
                    if v not in values:
                        print("Error: Flagged option {} not in question options. Question {}".format(v, question))
                    else:
                        flagged.append(v)

            if more_option:
                more_option = more_option.strip().lower()
                if more_option not in values:
                    print("Error: More option {} not in question options. Question {}".format(more_option, question))

            for v in values:
                vkey = v
                free_text = 0
                if more_option and more_option == v:
                    vkey = '*' + v
                    free_text = 1

                try:
                    opt_id = options[vkey]
                except KeyError:
                    option_id += 1
                    opt_id = option_id
                    out.write("INSERT INTO Options (Id, IsFreeText, Text, Hint) "
                              + "VALUES ({}, {}, '{}', null);\n".format(opt_id, free_text, v))
                    options[vkey] = option_id

                o2o_id += 1
                out.write("INSERT INTO OptionsToQuestions (Id, IdQuestion, IdOption, Flagged) "
                          + "VALUES ({}, {}, {}, {});\n".format(o2o_id, question_id, opt_id, 1 if v in flagged else 0))
                out.flush()

        else:
            if question_type == TYPES['Single'] or question_type == TYPES['Multiple']:
                print("Error: Missing values for question {}".format(question))
            else:
                o2o_id += 1
                out.write("INSERT INTO OptionsToQuestions (Id, IdQuestion, IdOption, Flagged) "
                          + "VALUES ({}, {}, {}, {});\n".format(o2o_id, question_id, FREE_OPTION_ID, 0))
                out.flush()

out.close()
